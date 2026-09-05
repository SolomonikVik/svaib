#!/usr/bin/env python3
"""read_metrics.py — чтение значений метрик из книги клиента по карте адресов.

Карта — эфемерный JSON, который агент выписывает, читая описание метрики в базе. В базе клиента
машинных артефактов не заводим: описание метрики остаётся единственным источником правды.

    python3 read_metrics.py --book <снимок>.xlsx --card <карта>.json [--upto 2026-09] [--json]

Что код выводит из книги сам: строку метрики (по меткам), ось периодов (шапка),
режим процента (формат ячейки). Что обязано быть в карте: книга, лист, путь меток,
единица и масштаб — из формата ячейки они не выводятся.

Расхождение книги с описанием метрики не отказ: строку сдвинули, лист переименовали, но метки
нашлись — читаем и говорим, насколько разошлось. Книга перестала опознаваться (метки не
находятся, шапка не парсится, кандидатов несколько) — отказ по этой метрике, без числа.
"""
from __future__ import annotations

import argparse, datetime as dt, json, re, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("нужен openpyxl: pip install openpyxl", file=sys.stderr); raise SystemExit(2)

MONTHS = {'jan':1,'january':1,'янв':1,'январь':1,'feb':2,'february':2,'фев':2,'февраль':2,
          'mar':3,'march':3,'мар':3,'март':3,'apr':4,'april':4,'апр':4,'апрель':4,
          'may':5,'май':5,'мая':5,'jun':6,'june':6,'июн':6,'июнь':6,'jul':7,'july':7,'июл':7,'июль':7,
          'aug':8,'august':8,'авг':8,'август':8,'sep':9,'sept':9,'september':9,'сен':9,'сентябрь':9,
          'oct':10,'october':10,'окт':10,'октябрь':10,'nov':11,'november':11,'ноя':11,'ноябрь':11,
          'dec':12,'december':12,'дек':12,'декабрь':12}


# Единицы, где ошибка в тысячу раз не видна глазом: «901» и «901 000» одинаково
# правдоподобны. Формат ячейки о масштабе молчит, поэтому для таких метрик карта
# обязана объявить `scale` — хоть 1, если числа в книге сырые.
MONEY_HINTS = ("aed", "usd", "eur", "gbp", "руб", "₽", "$", "€", "£", "тыс", "млн",
               "money", "amount", "выручк", "маржа", "прибыл", "расход")


def needs_scale(unit):
    u = norm(unit)
    if not u:
        return False
    if u in ("%", "percent", "процент", "шт", "штук", "pcs", "клиенты", "человек", "дней", "days"):
        return False
    return any(h in u for h in MONEY_HINTS)


def norm(s):
    return '' if s is None else re.sub(r'\s+', ' ', str(s)).strip().lower()


def parse_month(v):
    k = norm(v).replace('.', '')
    if not k:
        return None
    if k in MONTHS:
        return MONTHS[k]
    for name, m in sorted(MONTHS.items(), key=lambda x: -len(x[0])):
        if k.startswith(name):
            return m
    return None


def sheet_grid(ws):
    """Значения листа с forward-fill по объединённым ячейкам: метка группы стоит один раз."""
    grid = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]
    for rng in ws.merged_cells.ranges:
        v = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                grid[r - 1][c - 1] = v
    return grid


def find_period_axis(grid, max_scan=10):
    """Шапка периодов: строка с наибольшим числом распознанных месяцев (минимум три)."""
    best = None
    for r in range(min(max_scan, len(grid))):
        cols = {}
        for c, v in enumerate(grid[r]):
            m = parse_month(v)
            if m and m not in cols:
                cols[m] = c
        if len(cols) >= 3 and (best is None or len(cols) > len(best[1])):
            best = (r, cols)
    return best


def _match(path, want, exact):
    i = 0
    for cell in path:
        if i < len(want) and want[i] and (cell == want[i] if exact else want[i] in cell):
            i += 1
    return i == len(want)


def find_rows(grid, labels, first_value_col):
    """Строки по пути меток. Сначала точное совпадение; подстрока — только если точных ноль.

    Подстрока на живых книгах ловит лишнее: «% churn rate» входит в «% churn rate partners»,
    «fact» — в «fact (2025)». Поэтому неоднозначность не разрешается: отдаём всех кандидатов,
    решение принимает вызывающий, а значение не выводится.
    """
    want = [norm(x) for x in labels]
    for exact in (True, False):
        hits = [r for r in range(len(grid))
                if _match([norm(grid[r][c]) for c in range(first_value_col)], want, exact)]
        if hits:
            return hits, ('exact' if exact else 'substring')
    return [], 'none'


def find_column(grid, label, max_scan=10):
    """Колонка по заголовку («Month», «Total») — для листов без ряда по месяцам."""
    want = norm(label)
    for r in range(min(max_scan, len(grid))):
        for c, v in enumerate(grid[r]):
            if norm(v) == want:
                return r, c
    return None, None


def read_by_column(grid, ws, item):
    """Значение на пересечении строки-по-меткам и колонки-по-заголовку.

    Так адресуются оперативные листы: ряда по месяцам нет, есть колонка «Month» и
    строки, различающиеся меткой блока («MACROREGION / Profitability forecast»).
    Координаты в карте не нужны — те же метки, что в описании метрики; строки съедут, а
    метки останутся.
    """
    hrow, col = find_column(grid, item["column_label"])
    if col is None:
        return None, f"колонка «{item['column_label']}» не найдена"
    hits, mode = find_rows(grid, item["labels"], col)
    if len(hits) != 1:
        return None, ("меток не найдено" if not hits
                      else f"неоднозначно: {len(hits)} строк {[h + 1 for h in hits]}")
    r = hits[0]
    v = grid[r][col]
    if v is None or not isinstance(v, (int, float)):
        return None, f"в строке {r + 1} нет числа"
    fmt = ws.cell(r + 1, col + 1).number_format
    val = round(v * 100, 2) if str(fmt).endswith('%') else v * (item.get("scale") or 1)
    return (val, r + 1, mode), None


def read_cell(ws, grid, item):
    """Адрес одной ячейкой: значение текущего периода, которого нет в ряду по месяцам.

    Так адресуются оперативные показатели («прогноз месяца», «накопленным итогом»):
    ряда по периодам у них нет, есть колонка «Month» и своя строка. Координата без
    проверки опасна — строки в таких листах съезжают, — поэтому карта обязана нести
    guard: ячейку с меткой и её ожидаемый текст. Не сошлось — отказ, не чтение соседа.
    """
    guard = item.get("guard") or {}
    if not guard.get("cell") or not guard.get("label"):
        return None, "адрес ячейкой без guard: нужна метка-подтверждение"
    got = ws[guard["cell"]].value
    if norm(got) != norm(guard["label"]):
        return None, (f"guard не сошёлся: {guard['cell']} = {got!r}, "
                      f"ожидалось {guard['label']!r} — строки съехали")
    cell = ws[item["cell"]]
    v = cell.value
    if v is None or not isinstance(v, (int, float)):
        return None, f"в {item['cell']} нет числа"
    return (round(v * 100, 2) if str(cell.number_format).endswith('%')
            else v * (item.get("scale") or 1)), None


def resolve_sheet(wb, name):
    """Лист по имени; переименовали — ищем единственный похожий и говорим об этом."""
    if name in wb.sheetnames:
        return name, None
    cand = [s for s in wb.sheetnames if norm(name) in norm(s) or norm(s) in norm(name)]
    if len(cand) == 1:
        return cand[0], f"лист «{name}» не найден, читаю «{cand[0]}»"
    return None, f"лист «{name}» не найден, похожих {len(cand)}"


def _read_impl(book, card, year=None, upto=None):
    """upto — последний период, за который бывает ФАКТ (по умолчанию текущий месяц).

    В книгах формулы считают и будущие месяцы: пустой сентябрь выдаёт 0, а рост
    год-к-году — 43%. Такой факт не факт, а артефакт формулы, поэтому ряд факта
    обрезается по upto. План на будущее остаётся: он и должен смотреть вперёд.
    """
    wb = openpyxl.load_workbook(book, data_only=True)
    cache, out = {}, []
    for item in card["metrics"]:
        row = {"metric": item["name"], "unit": item.get("unit"), "notes": [], "values": {}, "plan": {},
               "companion": bool(item.get("year_offset"))}
        if needs_scale(item.get("unit")) and item.get("scale") is None:
            # Читаем как есть, но вслух: масштаб денежной метрики глазом не проверяется,
            # «901» и «901 000» одинаково правдоподобны. Молчать здесь нельзя, запрещать —
            # тоже: у метрики может не быть описания вовсе, а число всё равно нужно.
            row["notes"].append(
                f"единица не подтверждена: не сказано, в чём числа — тысячи или сырые "
                f"«{item.get('unit')}». Числа отданы как в книге")
            row["gap"] = {"what": "unit", "metric": item["name"],
                          "unit": item.get("unit"), "spec_file": item.get("spec_file")}

        sheet, note = resolve_sheet(wb, item["sheet"])
        if note:
            row["notes"].append(note)
        if sheet is None:
            row["status"] = "refused"; row["reason"] = note; out.append(row); continue
        if sheet not in cache:
            ws = wb[sheet]; g = sheet_grid(ws); cache[sheet] = (ws, g, find_period_axis(g))
        ws, grid, axis = cache[sheet]

        if item.get("column_label"):
            got, err = read_by_column(grid, ws, item)
            if err:
                row["status"] = "refused"; row["reason"] = err; out.append(row); continue
            value, r, mode = got
            row["row"] = r
            period = item.get("period") or upto or "текущий"
            row["values"] = {period: value}
            row["notes"].append(item.get("note") or "оперативное значение, период не закрыт")
            if mode == 'substring':
                row["notes"].append("метка найдена неточно")
            row["status"] = "ok_with_notes"; out.append(row); continue

        if item.get("cell"):
            value, err = read_cell(ws, grid, item)
            if err:
                row["status"] = "refused"; row["reason"] = err; out.append(row); continue
            period = item.get("period") or upto or ""
            row["values"] = {period: value} if period else {"текущий": value}
            row["notes"].append(item.get("note") or "оперативное значение, период не закрыт")
            row["status"] = "ok_with_notes"; out.append(row); continue

        if not axis:
            row["status"] = "refused"; row["reason"] = "ось периодов не распознана"; out.append(row); continue
        _, cols = axis
        first = min(cols.values())

        def series(labels, kind):
            hits, mode = find_rows(grid, labels, first)
            if len(hits) != 1:
                return None, (f"меток не найдено" if not hits
                              else f"неоднозначно: {len(hits)} строк {[h + 1 for h in hits]}"), None
            r = hits[0]
            fmt = ws.cell(r + 1, first + 1).number_format
            pct = fmt.endswith('%')
            scale = item.get("scale") or 1
            vals = {}
            for m, c in sorted(cols.items()):
                v = grid[r][c]
                if v is None or not isinstance(v, (int, float)):
                    continue
                base_year = (year or card.get('year') or 0) + (item.get("year_offset") or 0)
                period = f"{base_year}-{m:02d}"
                if kind == "факт" and upto and not item.get("year_offset"):
                    if period > upto:
                        continue
                    # Текущий (незакрытый) месяц: ноль здесь означает «ещё не заполнено»,
                    # а не «ноль продаж» — в прошлых месяцах ноль остаётся значением.
                    if period == upto and v == 0:
                        continue
                vals[period] = round(v * 100, 2) if pct else v * scale
            if mode == 'substring':
                row["notes"].append(f"{kind}: метка найдена неточно")
            return r + 1, None, vals

        r, err, vals = series(item["labels"], "факт")
        if err:
            row["status"] = "refused"; row["reason"] = err; out.append(row); continue
        row["row"] = r; row["values"] = vals
        now = dt.date.today().strftime("%Y-%m")
        if now in vals:
            row["notes"].append(f"{now} — месяц не закрыт, значение неокончательное")
        if item.get("expected_row") and item["expected_row"] != r:
            row["notes"].append(f"строка была {item['expected_row']}, стала {r} — книга разошлась с описанием метрики")
        if item.get("plan_labels"):
            pr, perr, pvals = series(item["plan_labels"], "план")
            if perr:
                row["notes"].append(f"план не прочитан: {perr}")
            else:
                row["plan"] = pvals
        row["status"] = "ok" if not row["notes"] else "ok_with_notes"
        out.append(row)
    return out


def read(book, card, year=None, upto=None):
    """Публичный вход: upto по умолчанию — текущий месяц."""
    return _read_impl(book, card, year or card.get("year"),
                      upto or dt.date.today().strftime("%Y-%m"))


def render(rows, source_note=None):
    """Таблица для агента-потребителя: метрика · период · план · факт · пометка."""
    lines = []
    if source_note:
        lines.append(source_note); lines.append("")
    for row in rows:
        head = f"**{row['metric']}**" + (f" ({row['unit']})" if row.get("unit") else "")
        if row["status"] == "refused":
            lines.append(f"{head}: значения нет — {row['reason']}"); lines.append(""); continue
        periods = sorted(row["values"])
        lines.append(head + (f" — {'; '.join(row['notes'])}" if row["notes"] else ""))
        lines.append("| период | " + " | ".join(periods) + " |")
        lines.append("|---" * (len(periods) + 1) + "|")
        lines.append("| факт | " + " | ".join(str(row["values"][p]) for p in periods) + " |")
        if row["plan"]:
            lines.append("| план | " + " | ".join(str(row["plan"].get(p, "—")) for p in periods) + " |")
        lines.append("")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--book", required=True, help="снимок книги (.xlsx)")
    ap.add_argument("--card", required=True, help="карта адресов из описаний метрик (.json)")
    ap.add_argument("--upto", help="последний период с фактом, YYYY-MM (по умолчанию текущий месяц)")
    ap.add_argument("--json", action="store_true", help="машинный вывод вместо таблицы")
    a = ap.parse_args()
    card = json.loads(Path(a.card).read_text(encoding="utf-8"))
    upto = a.upto or dt.date.today().strftime("%Y-%m")
    rows = read(a.book, card, card.get("year") or dt.date.today().year, upto)
    if a.json:
        print(json.dumps(rows, ensure_ascii=False, indent=2))
    else:
        print(render(rows, card.get("source_note")))
    return 0 if any(r["status"] != "refused" for r in rows) else 1


if __name__ == "__main__":
    raise SystemExit(main())
