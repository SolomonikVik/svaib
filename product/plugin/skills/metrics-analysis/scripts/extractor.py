#!/usr/bin/env python3
"""Extractor — actual-сторона: физическое чтение snapshot'а по декларативной раскладке.

Раскладка (`extractors/*.json`, схема extractor-layout) остаётся per-source и
захардкоженной, как требует канон; меняется только носитель — данные вместо кода.
Универсального extractor'а это не вводит: нет раскладки под адрес → источник получает
status = no_extractor.

Три guard-а, любой останавливает чтение ВСЕГО источника (status = schema_mismatch,
метрики → not_read / blocked / mismatch, прогон продолжается с rc 0):
  1. schema_hash не сошёлся — структуру источника перестроили;
  2. метка-подтверждение на координате не совпала — строка съехала;
  3. тип значения разошёлся с ожидаемым — сменили формат числа.

Сеть не используется: читается только то, что лежит в --snapshot-dir.
"""
from __future__ import annotations

import collections
import datetime as dt
import hashlib
import json
import re
from pathlib import Path

from source_map import (BaseConfigError, canonical_source, config_files, normalize_label,
                        normalize_labels, _load_json, _validate)

EXTRACTORS_SUBDIR = "extractors"

# Кэшированные ошибки формул Excel: div0 отделён от прочих (extractor.md — enum статусов).
_DIV0 = "#DIV/0!"
_ERROR_CELLS = ("#VALUE!", "#REF!", "#N/A", "#NAME?", "#NUM!", "#NULL!", "#GETTING_DATA")

_MONTHS = {
    "jan": 1, "january": 1, "янв": 1, "январь": 1,
    "feb": 2, "february": 2, "фев": 2, "февраль": 2,
    "mar": 3, "march": 3, "мар": 3, "март": 3,
    "apr": 4, "april": 4, "апр": 4, "апрель": 4,
    "may": 5, "май": 5, "мая": 5,
    "jun": 6, "june": 6, "июн": 6, "июнь": 6,
    "jul": 7, "july": 7, "июл": 7, "июль": 7,
    "aug": 8, "august": 8, "авг": 8, "август": 8,
    "sep": 9, "sept": 9, "september": 9, "сен": 9, "сентябрь": 9,
    "oct": 10, "october": 10, "окт": 10, "октябрь": 10,
    "nov": 11, "november": 11, "ноя": 11, "ноябрь": 11,
    "dec": 12, "december": 12, "дек": 12, "декабрь": 12,
}

_ISO_MONTH = re.compile(r"^(\d{4})-(0[1-9]|1[0-2])$")
_ISO_WEEK = re.compile(r"^(\d{4})-w(0[1-9]|[1-4]\d|5[0-3])$")
_SHORT_WEEK = re.compile(r"^w(0?[1-9]|[1-4]\d|5[0-3])$")
_ISO_QUARTER = re.compile(r"^(\d{4})-q([1-4])$")
_SHORT_QUARTER = re.compile(r"^q([1-4])$")
_YEAR = re.compile(r"^(\d{4})$")

# Форматы заголовка периода, объявляемые раскладкой явно. Общий разбор их не
# знает намеренно: «01.26» при месячной шкале — январь 2026, при квартальной
# осмысленного чтения нет вовсе, а в третьей книге те же цифры могут означать
# дату. Угадывать здесь нельзя — можно только следовать объявлению.
_FMT_MM_YY = re.compile(r"^(0[1-9]|1[0-2])\.(\d{2})$")
_FMT_MM_YYYY = re.compile(r"^(0[1-9]|1[0-2])\.(\d{4})$")
_FMT_YYYY_MM = re.compile(r"^(\d{4})-(0[1-9]|1[0-2])$")
_FMT_ISO_DATE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")

#: Шкалы, с которыми формат сочетается. `iso-date` несёт полную дату и потому
#: сворачивается в любую шкалу; месячные форматы — только в месячную.
PERIOD_FORMAT_GRANULARITY = {
    "mm.yy": ("month",),
    "mm.yyyy": ("month",),
    "yyyy-mm": ("month",),
    "iso-date": ("month", "quarter", "week", "year"),
}

#: Канонические роли колонки. Читается ровно одна — `fact`; остальные объявляются,
#: чтобы колонка факта находилась однозначно, а не «первой слева».
CANONICAL_ROLES = ("fact", "plan", "forecast", "delta")
READ_ROLE = "fact"


#: Результат чтения координаты. Кортеж стал именованным, когда к статусу и
#: значению добавились заголовок прочитанной колонки (его сверяет verifier) и
#: причина недостачи (она уходит человеку в `source_ref`): позиционная распаковка
#: шести полей читалась бы как загадка.
ValueRead = collections.namedtuple(
    "ValueRead", "status value labels period column_label detail")


class SnapshotUnreadable(Exception):
    """Файл snapshot есть и sha256 сошёлся, но прочитать его нельзя (деградация, rc 0)."""


class GuardStop(Exception):
    """Guard остановил чтение источника: структура разошлась с ожидаемой."""

    def __init__(self, detail):
        super().__init__(detail)
        self.detail = detail


def parse_period_formatted(header, granularity, period_format):
    """Заголовок → период по ОБЪЯВЛЕННОМУ формату; None — колонка не период.

    Строгая замена общего разбора, а не дополнение к нему, и ветка дат сюда не
    попадает намеренно: «01.26» в xlsx вполне может лежать датой 2026-01-01 с
    форматом отображения `MM.YY`, а годовой итог — датой 2026-01-01 с форматом
    `YYYY`. Общий разбор прочитал бы вторую колонку январём — то есть отдал бы
    годовое число как месячный факт.
    """
    if isinstance(header, (dt.datetime, dt.date)) or header is None:
        return None
    text = normalize_label(header)
    if not text:
        return None

    if period_format == "mm.yy":
        m = _FMT_MM_YY.match(text)
        # Двузначный год — всегда 2000-е: книг с данными прошлого века у нас нет,
        # а «26» без правила читается и как 1926.
        return "20{0}-{1}".format(m.group(2), m.group(1)) if m else None
    if period_format == "mm.yyyy":
        m = _FMT_MM_YYYY.match(text)
        return "{0}-{1}".format(m.group(2), m.group(1)) if m else None
    if period_format == "yyyy-mm":
        m = _FMT_YYYY_MM.match(text)
        return "{0}-{1}".format(m.group(1), m.group(2)) if m else None
    if period_format == "iso-date":
        m = _FMT_ISO_DATE.match(text)
        if not m:
            return None
        try:
            day = dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
        return _period_of_date(day, granularity)
    return None


def _period_of_date(day, granularity):
    """Дата → период шкалы. Одна свёртка на обе ветки разбора."""
    if granularity == "month":
        return "{0:04d}-{1:02d}".format(day.year, day.month)
    if granularity == "quarter":
        return "{0:04d}-Q{1}".format(day.year, (day.month - 1) // 3 + 1)
    if granularity == "week":
        year, week, _ = day.isocalendar()
        return "{0:04d}-W{1:02d}".format(year, week)
    return "{0:04d}".format(day.year)


def parse_period(header, granularity, period_year, period_format=None):
    """Заголовок колонки → период в шкале листа; None — колонка не является периодом.

    Нераспознанный заголовок значением не считается: молчаливый сдвиг на соседнюю
    колонку — это дефект A1.2, где ряд остаётся правдоподобным, а период чужой.
    """
    if period_format:
        return parse_period_formatted(header, granularity, period_format)
    if header is None:
        return None
    if isinstance(header, (dt.datetime, dt.date)):
        return _period_of_date(header.date() if isinstance(header, dt.datetime) else header,
                               granularity)

    text = normalize_label(header).replace(".", " ").replace("'", " ").strip()
    if not text:
        return None
    compact = text.replace(" ", "")

    if granularity == "month":
        m = _ISO_MONTH.match(compact)
        if m:
            return f"{m.group(1)}-{m.group(2)}"
        word = compact.rstrip(",")
        month = _MONTHS.get(word)
        if month is None:
            # «jan 2026» / «январь 2026»
            parts = text.split()
            if len(parts) == 2 and parts[0] in _MONTHS and _YEAR.match(parts[1]):
                return f"{int(parts[1]):04d}-{_MONTHS[parts[0]]:02d}"
            return None
        if period_year is None:
            return None
        return f"{period_year:04d}-{month:02d}"
    if granularity == "week":
        m = _ISO_WEEK.match(compact)
        if m:
            return f"{m.group(1)}-W{m.group(2)}"
        m = _SHORT_WEEK.match(compact)
        if m and period_year is not None:
            # Только 1–53: «w0» дало бы период 2026-W00, который замороженная схема
            # отвергает — rc 0 при отчёте, не проходящем контракт.
            return f"{period_year:04d}-W{int(m.group(1)):02d}"
        return None
    if granularity == "quarter":
        m = _ISO_QUARTER.match(compact)
        if m:
            return f"{m.group(1)}-Q{m.group(2)}"
        m = _SHORT_QUARTER.match(compact)
        if m and period_year is not None:
            return f"{period_year:04d}-Q{m.group(1)}"
        return None
    # Год ищется в тексте С РАЗДЕЛИТЕЛЯМИ, а не в склейке: «01.26» после замены
    # точки на пробел даёт «01 26», и склейка «0126» проходила бы за год. Такое
    # значение уезжало в отчёт периодом «0126» — четыре цифры, замороженную схему
    # проходит, валидатор молчит. Заголовки вида «jan. 2026» правило не задевает:
    # их разбирает месячная ветка выше.
    m = _YEAR.match(text)
    return m.group(1) if m else None


class Sheet:
    """Прочитанный лист: сетка значений с координатами 1-based, как в xlsx."""

    def __init__(self, title, rows):
        self.title = title
        self._rows = rows

    def cell(self, row, col):
        if row < 1 or row > len(self._rows):
            return None
        line = self._rows[row - 1]
        if col < 1 or col > len(line):
            return None
        return line[col - 1]

    @property
    def max_row(self):
        return len(self._rows)

    @property
    def max_col(self):
        return max((len(r) for r in self._rows), default=0)


def load_snapshot(path):
    """Книга snapshot'а → {имя листа: Sheet}. Формат — по расширению файла."""
    path = Path(path)
    if path.suffix == ".xlsx":
        try:
            import openpyxl

            book = openpyxl.load_workbook(path, data_only=True)
        except Exception as exc:  # openpyxl бросает разное на битых архивах
            raise SnapshotUnreadable(f"xlsx не открывается: {exc}") from exc
        sheets = {}
        for worksheet in book.worksheets:
            rows = [list(r) for r in worksheet.iter_rows(values_only=True)]
            sheets[worksheet.title] = Sheet(worksheet.title, rows)
        return sheets
    if path.suffix == ".json":
        # json-слепок — только фикстуры и тесты; боевой формат всегда xlsx.
        try:
            doc = json.loads(path.read_text(encoding="utf-8"))
            raw = doc["sheets"]
            if not isinstance(raw, dict):
                raise TypeError("sheets должен быть объектом «имя листа → строки»")
            sheets = {}
            for name, rows in raw.items():
                sheets[name] = Sheet(name, [list(r) for r in rows])
        except (OSError, ValueError, KeyError, TypeError, AttributeError,
                RecursionError) as exc:
            # Любая кривизна слепка — деградация источника, а не падение прогона:
            # rc остаётся 0, книга получает snapshot_unreadable.
            raise SnapshotUnreadable(f"json-слепок не читается: {exc}") from exc
        return sheets
    raise SnapshotUnreadable(f"неизвестное расширение snapshot'а: {path.suffix}")


def load_layouts(base):
    """Раскладки базы: {адрес источника: layout}. Дубль адреса — отказ, не выбор.

    Отказ **отложенный**: конфликт запоминается и срывает прогон только когда конфликтная
    книга действительно понадобилась. Иначе забытая копия раскладки постороннего источника
    (`weekly-copy.json`, оставшаяся в базе клиента) роняла бы автоматический прогон повестки,
    ничего не решая по существу — а до канонизации адресов такая пара вообще мирно
    сосуществовала как раскладки «разных» книг.
    """
    layouts = {}
    origin = {}
    conflicts = {}
    unsupported = {}
    for path in config_files(base, subdir=EXTRACTORS_SUBDIR):
        doc = _load_json(path)
        newer = _layout_is_newer_than_engine(doc)
        if newer:
            # Версия проверяется ДО схемы: `additionalProperties: false` отвергнет
            # незнакомые поля раньше, чем дело дойдёт до разбора версии, и вместо
            # понятного «раскладка новее движка» получился бы отказ формы.
            raw_source = (doc.get("source") or "").strip()
            if not raw_source:
                raise BaseConfigError(
                    f"{path.name}: раскладка версии {newer} новее движка и не называет источник — "
                    "отложить отказ не к чему"
                )
            # Раскладка будущей версии — это «под книгу нет читаемой раскладки»,
            # то есть штатный no_extractor по ОДНОМУ источнику. Отказ прогона был
            # бы хуже: один файл из будущего лишил бы повестки все книги клиента,
            # включая те, что движок читает без вопросов.
            unsupported[canonical_source(raw_source)] = (
                f"{path.name}: раскладка версии {newer} новее движка "
                f"(поддерживается 1.{SUPPORTED_LAYOUT_MINOR}) — обновите образ"
            )
            continue
        _validate(doc, "extractor-layout.schema.json", path)
        source = canonical_source(doc["source"])
        if source in layouts:
            conflicts[source] = (
                f"источник {source!r} описан двумя раскладками: "
                f"{origin[source]} и {path.as_posix()} — какая верна, код решать не вправе"
            )
            continue
        for sheet_name, sheet_layout in doc["sheets"].items():
            if sheet_layout.get("percent_mode") and sheet_layout.get("value_scale") is not None:
                # Оба на листе означали бы ×100 и множитель разом — на смешанном листе
                # это тихая ошибка на два-три порядка. Масштаб задаётся построчно.
                raise BaseConfigError(
                    f"{path.name}: у листа {sheet_name!r} заданы и percent_mode, и "
                    "value_scale — на смешанном листе масштаб указывается на строке"
                )
            _check_sheet_roles(path, sheet_name, sheet_layout)
            _check_sheet_period_format(path, sheet_name, sheet_layout)
        seen_metrics = set()
        for entry in doc["rows"]:
            key = (entry["metric"]["name"], entry["metric"]["file"])
            if key in seen_metrics:
                raise BaseConfigError(
                    f"{path.name}: метрика {key} описана двумя строками раскладки — "
                    "какая верна, код решать не вправе"
                )
            seen_metrics.add(key)
            if entry["sheet"] not in doc["sheets"]:
                raise BaseConfigError(
                    f"{path.name}: строка ссылается на лист {entry['sheet']!r}, "
                    "которого нет в sheets"
                )
            first, last = doc["sheets"][entry["sheet"]]["block_rows"]
            if not first <= entry["row"] <= last:
                # Иначе значение читалось бы, но координата не входила бы ни в
                # schema_hash, ни в скан сирот — сверка источника стала бы неполной
                # при формально валидном отчёте.
                raise BaseConfigError(
                    f"{path.name}: строка {entry['row']} листа {entry['sheet']!r} вне "
                    f"объявленного блока [{first}, {last}] — она не попала бы ни в "
                    "отпечаток структуры, ни в перечень сирот"
                )
        layouts[source] = doc
        origin[source] = path.as_posix()
    return Layouts(layouts, conflicts, unsupported)


class Layouts(dict):
    """Раскладки базы плюс отложенные конфликты адресов.

    Ведёт себя как обычный словарь `{адрес: layout}`, но обращение к конфликтной книге —
    отказ: выбирать между двумя раскладками одной книги код не вправе, а молчать нельзя,
    иначе какая из них применится к снимку, решал бы порядок обхода каталога.
    """

    def __init__(self, layouts, conflicts, unsupported=None):
        super().__init__(layouts)
        self.conflicts = dict(conflicts)
        # Раскладки, которые физически есть, но новее движка: источник получает
        # штатный `no_extractor`, а причина остаётся здесь — чтобы «раскладки
        # нет» и «раскладка из будущего» различал хотя бы оператор в логе.
        self.unsupported = dict(unsupported or {})

    def _guard(self, source):
        if source in self.conflicts:
            raise BaseConfigError(self.conflicts[source])

    def __contains__(self, source):
        self._guard(source)
        return super().__contains__(source)

    def get(self, source, default=None):
        self._guard(source)
        return super().get(source, default)

    def __getitem__(self, source):
        self._guard(source)
        return super().__getitem__(source)


def _labels_at(sheet, layout_sheet, row):
    """Метки строки после forward-fill служебных колонок внутри блока.

    Forward-fill обязателен: в живых книгах блок назван один раз (merged-ячейка),
    у продолжений служебная колонка пуста, и без протяжки строка «fact (total all)»
    недостижима.
    """
    first_row = layout_sheet["block_rows"][0]
    labels = []
    for col in layout_sheet["label_columns"]:
        value = None
        for r in range(first_row, row + 1):
            cell = sheet.cell(r, col)
            if cell is not None and str(cell).strip():
                value = cell
        labels.append(value)
    return labels


#: Максимальная minor-версия раскладки, которую понимает этот движок. Major
#: фиксирован схемой (`^1\.`), поэтому сравнивается только minor.
SUPPORTED_LAYOUT_MINOR = 2


def _layout_is_newer_than_engine(doc):
    """Версия раскладки, если она новее движка; иначе None.

    Раскладка кладётся в базу клиента, а движок едет образом — и порядок бывает
    нарушен. Разница должна называться разницей версий, а не отказом формы:
    иначе оператор ищет опечатку в файле, который просто новее.
    """
    if not isinstance(doc, dict):
        # Не объект — это дефект формы, а не версии: пусть его назовёт схема,
        # иначе отказ придёт как usage_error вместо «раскладка не читается».
        return None
    contract = doc.get("contract")
    version = ((contract if isinstance(contract, dict) else {}).get("version") or "").strip()
    parts = version.split(".")
    if len(parts) != 3:
        return None
    try:
        major, minor = int(parts[0]), int(parts[1])
    except ValueError:
        return None
    if major == 1 and minor > SUPPORTED_LAYOUT_MINOR:
        return version
    return None


def _check_sheet_roles(path, sheet_name, layout_sheet):
    """Роли колонок: объявление либо полное, либо отсутствует.

    Проверки живут здесь, а не только в JSON-схеме, потому что схема без
    установленного `jsonschema` деградирует в no-op — и битая карта ролей прошла
    бы именно там, где страховка нужна. Каждый отказ — конфигурация базы, а не
    данные: чинится правкой раскладки, и выглядеть отсутствием данных не должен.
    """
    where = f"{path.name}: лист {sheet_name!r}"
    role_row = layout_sheet.get("role_row")
    roles = layout_sheet.get("roles")
    if role_row is None and roles is None:
        return
    if role_row is None or roles is None:
        raise BaseConfigError(
            f"{where}: role_row и roles задаются только вместе — половина объявления "
            "означала бы, что роль колонки то читается, то нет"
        )
    if not isinstance(roles, dict) or not roles:
        raise BaseConfigError(f"{where}: roles пуст — читать факт неоткуда")

    index = {}
    for text, role in roles.items():
        if role not in CANONICAL_ROLES:
            raise BaseConfigError(
                f"{where}: роль {role!r} для колонки {text!r} не из перечня "
                f"{', '.join(CANONICAL_ROLES)} — опечатка молча убрала бы колонку из оси"
            )
        key = normalize_label(text)
        if not key:
            # Пустой ключ после нормализации объявил бы ролью ПУСТУЮ ячейку
            # шапки — ровно то, что запрещает правило «пустая роль → колонка
            # выпадает», только чёрным ходом.
            raise BaseConfigError(
                f"{where}: заголовок {text!r} после нормализации пуст — такой ключ объявил бы "
                "ролью пустую ячейку шапки"
            )
        known = index.setdefault(key, role)
        if known != role:
            raise BaseConfigError(
                f"{where}: заголовки {text!r} и другой ключ совпадают после нормализации, "
                f"но объявлены разными ролями ({known} и {role})"
            )
    if READ_ROLE not in index.values():
        raise BaseConfigError(
            f"{where}: ни один заголовок не объявлен ролью {READ_ROLE!r} — "
            "значения читать неоткуда"
        )


def _check_sheet_period_format(path, sheet_name, layout_sheet):
    """Формат заголовка периода: сочетаемость со шкалой и с period_year."""
    period_format = layout_sheet.get("period_format")
    if period_format is None:
        return
    where = f"{path.name}: лист {sheet_name!r}"
    allowed = PERIOD_FORMAT_GRANULARITY.get(period_format)
    if allowed is None:
        raise BaseConfigError(f"{where}: неизвестный period_format {period_format!r}")
    granularity = layout_sheet.get("granularity")
    if granularity not in allowed:
        raise BaseConfigError(
            f"{where}: period_format {period_format!r} несовместим со шкалой {granularity!r} "
            f"(допустимо: {', '.join(allowed)})"
        )
    if layout_sheet.get("period_year") is not None:
        raise BaseConfigError(
            f"{where}: period_year вместе с period_format — два источника года, "
            "и какой побеждает, из раскладки не следует"
        )


def roles_index(layout_sheet):
    """{нормализованный текст шапки: каноническая роль}.

    Нормализация та же, что у меток строк: лишний пробел в шапке книги иначе
    убирает роль `fact` из оси, и все метрики листа становятся пустыми при
    исправных данных.
    """
    return {normalize_label(text): role for text, role in (layout_sheet.get("roles") or {}).items()}


def _column_role(sheet, layout_sheet, col, index):
    """Роль колонки; None — роль не объявлена или не распознана.

    Без `role_row` роль колонки одна по построению: книга держит один столбец на
    период, и он же факт. Forward-fill по ролям не делается намеренно — протяжка
    объявила бы соседнюю колонку фактом.
    """
    role_row = layout_sheet.get("role_row")
    if not role_row:
        return READ_ROLE
    return index.get(normalize_label(sheet.cell(role_row, col)))


def _scan_columns(sheet, layout_sheet):
    """[(период, роль, колонка)] по объявленной оси листа."""
    index = roles_index(layout_sheet)
    found = []
    for col in range(layout_sheet["value_columns_from"], sheet.max_col + 1):
        period = parse_period(
            sheet.cell(layout_sheet["period_row"], col),
            layout_sheet["granularity"],
            layout_sheet.get("period_year"),
            layout_sheet.get("period_format"),
        )
        if period is None:
            continue
        found.append((period, _column_role(sheet, layout_sheet, col, index), col))
    return found


def _period_axis(sheet, layout_sheet):
    """Периоды листа НЕЗАВИСИМО от ролей.

    На этой оси стоит ответ «период в книге есть, а колонки факта за него нет».
    Считать её по парам (период, роль) нельзя: если клиент переименовал роли всех
    колонок периода, пар не возникает вовсе — и недостача факта выглядела бы
    отсутствием данных, то есть ровно тем, что здесь и различается.
    """
    return {period for period, _role, _col in _scan_columns(sheet, layout_sheet)}


def _period_columns(sheet, layout_sheet):
    """{период: колонка} для читаемой роли — факта.

    Повтор периода в факте — неоднозначность, а не повод взять левую колонку:
    две колонки «2026-07 / факт» это обычно старая таблица рядом с актуальной, и
    «первая слева» молча отдаёт прошлогоднее число как свежее. Такой лист
    останавливается guard-ом.

    Дубль нечитаемой роли guard-ом не является: два прогнозных столбца под одним
    месяцем — законная форма живой книги, а читать их всё равно никто не будет.
    Поэтому функция знает ровно одну роль: параметр «какую читать» означал бы,
    что где-то читается не факт, а такого пути нет.
    """
    columns = {}
    for period, column_role, col in _scan_columns(sheet, layout_sheet):
        if column_role != READ_ROLE:
            continue
        if period in columns:
            raise GuardStop(
                f"на листе «{sheet.title}» период {period} встречается дважды "
                f"(колонки {columns[period]} и {col}) — какая актуальна, код решать не вправе"
            )
        columns[period] = col
    return columns


def _value_columns(sheet, layout_sheet):
    """Колонки с распознанным периодом И распознанной ролью — для инвентаризации.

    Фильтр по `fact` здесь неверен: строка, заполненная одним лишь планом, — всё
    ещё заполненная строка книги, и прятать её из перечня сирот нельзя.

    Guard дубля проверяется и здесь: неоднозначность оси — свойство листа, а не
    запроса, и лист с двумя колонками одного факта обязан останавливаться, даже
    если запрошенные метрики читаются с соседнего листа.
    """
    _period_columns(sheet, layout_sheet)
    return [col for _period, role, col in _scan_columns(sheet, layout_sheet) if role is not None]


def schema_hash(sheets, layout):
    """Отпечаток структуры: только листы раскладки.

    Состав: имя листа, заголовки колонок его period_row, метки служебных колонок всех
    строк блока. Метки в отпечатке — то, что делает исполнимым guard «сдвиг строки
    останавливает прогон»: канонический скелет из листов и заголовков сдвига не видит.
    Ограничение листами раскладки — осознанное отступление (runner-spec §4.3): хэш по
    всей книге давал бы schema_mismatch на каждую правку постороннего листа рабочего
    дашборда клиента.
    """
    skeleton = []
    for name in sorted(layout["sheets"]):
        layout_sheet = layout["sheets"][name]
        sheet = sheets.get(name)
        if sheet is None:
            skeleton.append({"sheet": name, "missing": True})
            continue
        headers = [
            normalize_label(sheet.cell(layout_sheet["period_row"], col))
            for col in range(1, sheet.max_col + 1)
        ]
        first, last = layout_sheet["block_rows"]
        labels = [
            list(normalize_labels(_labels_at(sheet, layout_sheet, row)))
            for row in range(first, min(last, sheet.max_row) + 1)
        ]
        skeleton.append({"sheet": name, "headers": headers, "labels": labels})
    digest = hashlib.sha256(
        json.dumps(skeleton, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()
    return f"sha256:{digest}"


def classify_cell(value):
    """Ячейка значения → (value_status, число).

    Три класса разведены намеренно — иначе один и тот же вход трактуется двояко:
      - число, в том числе записанное текстом («1 029», «6,8») → ok;
      - дата, время, булево → error: негодный ТИП значения одной метрики. Дата в xlsx —
        число-serial, без явного отказа она уехала бы правдоподобным пятизначным числом;
        True в числовом контексте молча стало бы единицей;
      - кэшированная ошибка формулы → div0 / error по enum extractor'а;
      - нечисловой текст («уточняется», подпись) → text_guard: значение не читается, а
        вызывающий останавливает источник целиком (guard 3) — подпись на месте числа
        означает, что таблицу перестроили, и это свойство структуры, а не одной ячейки.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return "missing", None
    if isinstance(value, bool):
        return "error", None
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return "error", None
    if isinstance(value, (int, float)):
        try:
            number = float(value)
        except (OverflowError, ValueError):
            # Целое за пределами float: значение прочитать нельзя, но это ошибка данных,
            # а не структуры — метрика деградирует, прогон идёт.
            return "error", None
        if number != number or number in (float("inf"), float("-inf")):
            # NaN и бесконечности — не JSON: сериализованные токенами NaN/Infinity они
            # сломали бы строгого потребителя, а как «значение» они бессмысленны.
            return "error", None
        return "ok", number
    text = str(value).strip()
    if text == _DIV0:
        return "div0", None
    if text in _ERROR_CELLS or (text.startswith("#") and text.endswith("!")):
        return "error", None
    cleaned = text.replace(" ", "").replace(" ", "").replace("%", "")
    # Запятая двусмысленна: «6,8» — десятичная дробь, «1,029» — тысячный разделитель,
    # и по одной ячейке их не различить. Догадка здесь стоила бы порядка величины,
    # поэтому неоднозначная форма (ровно три цифры после запятой) уходит в guard,
    # а не в число.
    if "," in cleaned and "." in cleaned:
        # «1.234,56» — европейская запись 1234.56, «1,234.56» — англоязычная. По одной
        # ячейке они неразличимы, а цена ошибки — порядок величины. Не гадаем.
        return "text_guard", None
    if cleaned.count(",") > 1:
        # «1,2,3» — не число: молча выбросив запятые, получили бы правдоподобное 123.
        return "text_guard", None
    if cleaned.count(",") == 1:
        left, right = cleaned.split(",")
        if len(right) == 3 and left.lstrip("-+").isdigit():
            # «12,500» — это и 12.5, и 12500. Тоже guard, а не догадка.
            return "text_guard", None
        cleaned = cleaned.replace(",", ".")
    try:
        number = float(cleaned)
    except (ValueError, OverflowError):
        return "text_guard", None
    if number != number or number in (float("inf"), float("-inf")):
        return "error", None
    return "ok", number


def apply_scale(value, layout_sheet, entry=None):
    """Приведение к единице паспорта: доля процента → проценты, затем множитель книги.

    Масштаб и режим процента задаются на листе и переопределяются на строке: живой лист
    weekly-метрик держит в одной таблице тысячи AED (MRR) и проценты (% of growth), и
    единый множитель на весь лист ошибся бы на три порядка в одну сторону либо на два
    в другую.
    """
    if value is None:
        return None
    entry = entry or {}
    percent_mode = entry.get("percent_mode", layout_sheet.get("percent_mode"))
    if percent_mode == "fraction":
        value = value * 100
    scale = entry.get("value_scale", layout_sheet.get("value_scale", 1))
    # Округление против двоичного шума: 0.084 * 100 даёт 8.399999999999998, и этот
    # хвост уехал бы руководителю. Девять знаков заведомо больше любой значащей
    # точности управленческой метрики.
    return round(value * scale, 9)


class SourceReading:
    """Результат чтения одного источника."""

    def __init__(self, source, status, error=None, schema_hash_value=None, detail=None):
        self.source = source
        self.status = status              # ok | unavailable | no_extractor | schema_mismatch
        self.error = error                # (code, message) | None
        self.schema_hash = schema_hash_value
        self.detail = detail              # текст для verification.schema_mismatch[]
        self.sheets = {}
        self.layout = None

    @property
    def readable(self):
        return self.status == "ok"


def read_source(source, layout, snapshot_path):
    """Открыть источник и прогнать guard-и уровня книги (hash).

    Guard-и уровня строки (метка, тип) исполняются при извлечении значения — но их
    срабатывание останавливает источник целиком, потому что «строка съехала» означает,
    что перестроена таблица, а не одна ячейка.
    """
    # Порядок проверок не произволен: если книги нет в манифесте, контракт требует
    # unavailable + snapshot_missing — независимо от того, есть ли под неё раскладка.
    # Обратный порядок дал бы no_extractor, и валидатор отверг бы такой отчёт.
    if snapshot_path is None:
        return SourceReading(
            source, "unavailable",
            error=("snapshot_missing", "книга объявлена паспортом, snapshot в манифесте отсутствует"),
        )
    if layout is None:
        return SourceReading(source, "no_extractor")
    try:
        sheets = load_snapshot(snapshot_path)
    except SnapshotUnreadable as exc:
        return SourceReading(source, "unavailable", error=("snapshot_unreadable", str(exc)))

    actual = schema_hash(sheets, layout)
    if actual != layout["schema_hash"]:
        reading = SourceReading(
            source, "schema_mismatch",
            error=("schema_mismatch", "структура книги разошлась с эталоном раскладки"),
            schema_hash_value=actual,
            detail=f"schema_hash разошёлся: эталон {layout['schema_hash']}, снимок {actual}",
        )
        reading.sheets, reading.layout = sheets, layout
        return reading

    reading = SourceReading(source, "ok", schema_hash_value=actual)
    reading.sheets, reading.layout = sheets, layout
    return reading


def assert_labels_unique(sheet, layout_sheet, entry):
    """Метки целевой строки не повторяются внутри блока — иначе координата ненадёжна.

    Схема не может потребовать, чтобы колонка plan/fact входила в `label_columns`:
    число служебных колонок зависит от книги. Но если её забыть, строки «MRR / plan»
    и «MRR / fact» дают ОДИНАКОВЫЙ кортеж меток — и перестановка строк проходит мимо
    и хэша, и метки-подтверждения, отдавая план как факт. Поэтому неоднозначность
    ловится структурно: одинаковые метки на читаемой координате — стоп.
    """
    target = normalize_labels(_labels_at(sheet, layout_sheet, entry["row"]))
    first, last = layout_sheet["block_rows"]
    columns = _value_columns(sheet, layout_sheet)
    for row in range(first, min(last, sheet.max_row) + 1):
        if row == entry["row"]:
            continue
        if _is_spacer_row(sheet, layout_sheet, row, columns):
            # Пустая строка-разделитель внутри блока: собственных меток нет, и
            # forward-fill протягивает на неё метки предыдущей строки. Двойником
            # она не является — читать из неё нечего, подменить она не может.
            # На живой книге таких строк шесть, и без этого условия guard
            # останавливал чтение всей книги при исправной раскладке.
            continue
        if normalize_labels(_labels_at(sheet, layout_sheet, row)) == target:
            raise GuardStop(
                f"на листе «{sheet.title}» строки {entry['row']} и {row} неразличимы по меткам "
                f"{[str(x) for x in entry['labels']]} — координата ненадёжна "
                "(не забыта ли колонка plan/fact в label_columns?)"
            )


def _is_spacer_row(sheet, layout_sheet, row, columns):
    """Строка-разделитель: ни собственной метки, ни единого значения.

    «Собственной» — то есть до forward-fill: именно протяжка делает такие строки
    неотличимыми от предыдущей. Строка со значениями разделителем не считается
    никогда, даже без меток: из неё можно прочитать число, значит она участвует
    в проверке уникальности координат.
    """
    for col in layout_sheet["label_columns"]:
        value = sheet.cell(row, col)
        if value is not None and str(value).strip():
            return False
    return all(classify_cell(sheet.cell(row, col))[0] == "missing" for col in columns)


def read_value(reading, entry, period, strict_guard=True):
    """Значение метрики по захардкоженной координате.

    Возвращает (value_status, число, фактические метки строки, фактический период).
    Метка-подтверждение проверяется ДО чтения значения: адрес — это координата плюс
    метка, потому что структуру источника правит клиент и строки съезжают.

    `strict_guard = False` — чтение ВСПОМОГАТЕЛЬНОЙ ячейки (база сравнения, кандидат
    fallback). Guard 3 привязан к ячейке ЗАПРОШЕННОГО периода: подпись «уточняется» в
    колонке, которую никто не спрашивал, не повод блокировать источник и потерять
    честно прочитанное число текущего периода.
    """
    layout_sheet = reading.layout["sheets"][entry["sheet"]]
    sheet = reading.sheets.get(entry["sheet"])
    if sheet is None:
        raise GuardStop(f"лист «{entry['sheet']}» отсутствует в книге")

    actual_labels = _labels_at(sheet, layout_sheet, entry["row"])
    if normalize_labels(actual_labels) != normalize_labels(entry["labels"]):
        raise GuardStop(
            f"метка строки {entry['row']} листа «{entry['sheet']}» не совпала: "
            f"ожидалось {entry['labels']}, найдено {[str(x) for x in actual_labels]}"
        )
    assert_labels_unique(sheet, layout_sheet, entry)

    columns = _period_columns(sheet, layout_sheet)
    if period not in columns:
        # Различаем два «нет значения»: периода нет в книге вовсе — и период есть,
        # а колонки факта за него нет (стоит прогноз или план). Второе выглядит
        # как отсутствие данных, хотя данные есть и они просто не факт.
        detail = None
        if period in _period_axis(sheet, layout_sheet):
            detail = f"{period}: в книге есть колонки этого периода, но нет колонки факта"
        return ValueRead("missing", None, actual_labels, None, None, detail)

    col = columns[period]
    column_label = _column_header(sheet, layout_sheet, col)
    raw = sheet.cell(entry["row"], col)
    status, value = classify_cell(raw)
    if status == "text_guard":
        if not strict_guard:
            return ValueRead("missing", None, actual_labels, None, column_label, None)
        raise GuardStop(
            f"в ячейке значения строки {entry['row']} листа «{entry['sheet']}» "
            f"нечисловой текст {str(raw)[:40]!r} — таблицу перестроили"
        )
    scaled = apply_scale(value, layout_sheet, entry)
    if scaled is not None and (scaled != scaled or scaled in (float("inf"), float("-inf"))):
        # Масштаб мог вынести значение за пределы float — тогда в отчёт уехал бы
        # нестрогий JSON-токен Infinity, который валидатор пропускает, а потребитель
        # разобрать не сможет.
        return ValueRead("error", None, actual_labels, period, column_label, None)
    return ValueRead(status, scaled, actual_labels, period, column_label, None)


def read_plan(reading, entry, period):
    """Плановое значение метрики — по отдельной координате раскладки.

    Возвращает число или None. Guard-и те же, что у факта: метка сверяется до
    чтения, масштаб применяется тот же. Отсутствие `plan_row` — не отказ, а
    штатное «плана в книге нет»: у производных метрик его не бывает вовсе.

    План читается ТОЛЬКО из источника. Плановое значение из цели сюда не
    подставляется: у них разные роли и разные периоды, и молчаливый выбор между
    ними канон прямо запрещает.
    """
    plan_row = entry.get("plan_row")
    if not plan_row:
        return None
    plan_entry = dict(entry, row=plan_row, labels=entry.get("plan_labels") or entry["labels"])
    try:
        read = read_value(reading, plan_entry, period, strict_guard=False)
    except GuardStop:
        # Съехавшая плановая строка не должна ронять факт: план — дополнение к
        # значению, а не оно само.
        return None
    return read.value if read.status == "ok" else None


def _column_header(sheet, layout_sheet, col):
    """Заголовок роли колонки как он написан в книге — сырой, не канон.

    Сверять `fact_column_label` карты нужно именно с ним: карта пишется человеком
    по книге, а не по нашему словарю ролей, и сравнение с каноническим `fact`
    отвергало бы верные карты.
    """
    role_row = layout_sheet.get("role_row")
    if not role_row:
        return None
    value = sheet.cell(role_row, col)
    return None if value is None else str(value).strip()


def available_periods(reading, entry):
    """Периоды, в которых у строки есть ЧИСЛО.

    Именно число, а не «непустая ячейка»: fallback ищет последнее известное значение,
    и `#DIV/0!` в предыдущей колонке не должен ни подменять его собой, ни закрывать
    дорогу к более старому числу. Ошибка значения за сам запрошенный период при этом
    показывается как есть — fallback включается только когда period пуст.
    """
    layout_sheet = reading.layout["sheets"][entry["sheet"]]
    sheet = reading.sheets[entry["sheet"]]
    found = []
    # Только колонки факта: fallback ищет последнее известное ЗНАЧЕНИЕ, и план
    # будущего месяца, подобранный «последним известным», уехал бы фактом.
    for period, col in _period_columns(sheet, layout_sheet).items():
        status, _ = classify_cell(sheet.cell(entry["row"], col))
        if status == "ok":
            found.append(period)
    return sorted(found)


def orphan_rows(reading, bound_rows, bound_labels):
    """Строки блоков, не названные ни одной привязкой карты источников.

    Исключаются, во-первых, КООРДИНАТЫ, которые раскладка реально читает (`bound_rows`),
    и только во-вторых — метки привязок, для которых координаты нет (`bound_labels`).
    Порядок важен: если в блоке две строки с меткой «MRR / fact» — старая таблица и
    актуальная, — исключение по метке спрятало бы обе, и подмена строки не всплыла бы
    даже сиротой. Исключение по координате оставляет двойника видимым.

    Кандидатом в сироты считается строка с непустой меткой И хотя бы одним непустым
    значением: строки-комментарии CEO между метриками (канон называет их отдельной
    патологией источников) сиротами не являются.
    Свёртка по (лист, метка) обязательна — иначе два одинаковых ряда дают два
    идентичных объекта и роняют весь отчёт по uniqueItems схемы.
    """
    found = {}
    for sheet_name, layout_sheet in reading.layout["sheets"].items():
        sheet = reading.sheets.get(sheet_name)
        if sheet is None:
            continue
        columns = _value_columns(sheet, layout_sheet)
        first, last = layout_sheet["block_rows"]
        for row in range(first, min(last, sheet.max_row) + 1):
            labels = _labels_at(sheet, layout_sheet, row)
            normalized = normalize_labels(labels)
            if not any(part for part in normalized):
                continue
            if (sheet_name, row) in bound_rows:
                continue
            if (normalize_label(sheet_name), normalized) in bound_labels:
                continue
            has_value = any(
                classify_cell(sheet.cell(row, col))[0] != "missing" for col in columns
            )
            if not has_value:
                continue
            key = (sheet_name, " / ".join(str(x) for x in labels if x is not None))
            found.setdefault(key, True)
    return sorted(found)
