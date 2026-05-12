"""
Probe an xlsx source during metrics onboarding.

This is a generic reconnaissance helper. It does not aggregate numbers and does
not decide metric semantics. It writes a markdown report with workbook shape,
sheet names, selected sheet dump, merged cells, formulas, cached formula errors,
and non-empty counts by column.

Usage:
    python3 probe_xlsx.py --xlsx metrics/source/source.xlsx --sheet "Sheet1" --out _findings/01_xlsx_structure.md
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


EXCEL_ERRORS = {"#DIV/0!", "#REF!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}


def cell_to_text(value: Any, max_len: int = 40) -> str:
    if pd.isna(value):
        return "."
    text = str(value)
    if len(text) > max_len:
        return text[: max_len - 1] + "..."
    return text


def write_report(xlsx: Path, sheet: str | None, out: Path, max_dump_rows: int | None) -> None:
    if not xlsx.exists():
        print(f"FAIL: xlsx not found: {xlsx}")
        sys.exit(1)

    out.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    lines.append("# XLSX structure probe")
    lines.append("")
    lines.append(f"File: `{xlsx.name}`")
    lines.append("")

    workbook = pd.ExcelFile(xlsx)
    lines.append("## Sheets")
    lines.append("")
    for index, name in enumerate(workbook.sheet_names, 1):
        lines.append(f"{index}. `{name}`")
    lines.append("")

    target_sheet = sheet or workbook.sheet_names[0]
    if target_sheet not in workbook.sheet_names:
        lines.append(f"## Error: sheet `{target_sheet}` not found")
        out.write_text("\n".join(lines), encoding="utf-8")
        print(f"FAIL: sheet not found: {target_sheet}")
        sys.exit(1)

    df_raw = pd.read_excel(xlsx, sheet_name=target_sheet, header=None)
    lines.append(f"## Sheet `{target_sheet}`")
    lines.append("")
    lines.append(f"Size: **{df_raw.shape[0]} x {df_raw.shape[1]}**")
    lines.append("")

    lines.append("### First rows")
    lines.append("")
    lines.append("```")
    for row_index in range(min(5, df_raw.shape[0])):
        row = [cell_to_text(value) for value in df_raw.iloc[row_index].tolist()]
        lines.append(f"row {row_index}: {row}")
    lines.append("```")
    lines.append("")

    row_limit = df_raw.shape[0] if max_dump_rows is None else min(max_dump_rows, df_raw.shape[0])
    lines.append("### Sheet dump")
    lines.append("")
    if row_limit < df_raw.shape[0]:
        lines.append(f"Showing first {row_limit} rows of {df_raw.shape[0]}.")
        lines.append("")
    lines.append("```")
    for row_index in range(row_limit):
        cells = [cell_to_text(value, max_len=30) for value in df_raw.iloc[row_index].tolist()]
        lines.append(f"r{row_index:03d}: " + " | ".join(cells))
    lines.append("```")
    lines.append("")

    wb_formula = load_workbook(xlsx, data_only=False)
    ws_formula = wb_formula[target_sheet]
    merged = list(ws_formula.merged_cells.ranges)
    formulas: list[tuple[str, str]] = []
    raw_errors: list[tuple[str, str]] = []
    for row in ws_formula.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str):
                if value.startswith("="):
                    formulas.append((cell.coordinate, value))
                if value in EXCEL_ERRORS:
                    raw_errors.append((cell.coordinate, value))

    wb_values = load_workbook(xlsx, data_only=True)
    ws_values = wb_values[target_sheet]
    cached_errors: list[tuple[str, str]] = []
    for row in ws_values.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value in EXCEL_ERRORS:
                cached_errors.append((cell.coordinate, value))

    lines.append("## Structural details")
    lines.append("")
    lines.append(f"- Merged cells: {len(merged)}")
    for merged_range in merged[:30]:
        lines.append(f"  - `{merged_range}`")
    if len(merged) > 30:
        lines.append(f"  - ... {len(merged) - 30} more")
    lines.append("")

    lines.append(f"- Formulas: {len(formulas)}")
    for coord, formula in formulas[:30]:
        lines.append(f"  - `{coord}` = `{formula}`")
    if len(formulas) > 30:
        lines.append(f"  - ... {len(formulas) - 30} more")
    lines.append("")

    lines.append(f"- Raw errors: {len(raw_errors)}")
    for coord, error in raw_errors[:30]:
        lines.append(f"  - `{coord}` = `{error}`")
    if len(raw_errors) > 30:
        lines.append(f"  - ... {len(raw_errors) - 30} more")
    lines.append("")

    lines.append(f"- Cached formula errors: {len(cached_errors)}")
    for coord, error in cached_errors[:30]:
        lines.append(f"  - `{coord}` = `{error}`")
    if len(cached_errors) > 30:
        lines.append(f"  - ... {len(cached_errors) - 30} more")
    lines.append("")

    lines.append("## Non-empty cells by column")
    lines.append("")
    lines.append("```")
    for col_index in range(df_raw.shape[1]):
        series = df_raw.iloc[:, col_index]
        non_empty = series.notna().sum()
        sample = [cell_to_text(value, max_len=25) for value in series.dropna().head(3).tolist()]
        lines.append(f"col {col_index:03d}: {non_empty:3d} non-empty | sample: {sample}")
    lines.append("```")
    lines.append("")

    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"OK: wrote {out}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Probe xlsx structure for metrics onboarding")
    parser.add_argument("--xlsx", required=True, type=Path, help="Path to xlsx file")
    parser.add_argument("--sheet", default=None, help="Sheet name. Defaults to first sheet")
    parser.add_argument("--out", required=True, type=Path, help="Markdown report path")
    parser.add_argument(
        "--max-dump-rows",
        type=int,
        default=None,
        help="Limit sheet dump rows. Defaults to full selected sheet.",
    )
    args = parser.parse_args()
    write_report(args.xlsx, args.sheet, args.out, args.max_dump_rows)


if __name__ == "__main__":
    main()
